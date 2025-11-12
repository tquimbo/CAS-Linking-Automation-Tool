# ===== In-place Qty writing (keeps formatting) =====
import re
import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl import load_workbook

# --- CONFIG: point to your actual file/sheet/headers ---
EXCEL_IN        = "test_budget.xlsx"        # <-- your exact file
EXCEL_SHEET     = "Sheet1"                  # tab name
HEADER_ROW_NUM  = 3                         # your headers are on row 3 (1-based)
ITEM_HEADER     = "Billing Description"     # exact header text in your sheet
QTY_HEADER      = "Qty"                     # exact header text in your sheet

# Matching thresholds
MIN_SCORE = 86
MIN_DELTA = 6

UNIT_STOPWORDS = set("""
per notice per email per record per letter per check per piece per trace per response per minute per month per mailing per form
one time one-time onetime minimum fee
packet packets notice notices email emails letter letters postcard postcards post card post cards
document documents claim claims web website portal phone fax faxes text texts
""".split())

def normalize_text(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = s.lower()
    s = s.replace("&", " and ")
    s = re.sub(r"\[[^\]]*\]", " ", s)   # remove [bracketed]
    s = re.sub(r"[^\w\s]", " ", s)      # remove punctuation
    s = re.sub(r"\s+", " ", s).strip()
    tokens = [t for t in s.split() if t not in UNIT_STOPWORDS]
    return " ".join(tokens)

def extract_numeric_qty(volume_str):
    """Pick first non-percent numeric token as Qty (fallback: any number)."""
    tokens = re.findall(r"[0-9][0-9,\.]*%?|%|[A-Za-z\-]+", str(volume_str))
    for t in tokens:
        t_clean = t.replace(",", "")
        if t_clean.endswith("%"):
            continue
        if re.fullmatch(r"\d+(\.\d+)?", t_clean):
            try:
                return int(float(t_clean)) if t_clean.isdigit() else float(t_clean)
            except:
                pass
    m = re.search(r"[\d,]+(?:\.\d+)?", str(volume_str).replace(",", ""))
    if m:
        val = m.group()
        try:
            return int(float(val)) if val.isdigit() else float(val)
        except:
            return None
    return None

def write_qty_in_place(parsed_df: pd.DataFrame):
    """
    parsed_df must have at least columns: Item, Qty (or Volume/VolumeRaw)
    Only writes Qty; preserves all Excel formatting; saves back to EXCEL_IN.
    """
    # Prepare PDF items
    pdf_items = []
    for _, row in parsed_df.iterrows():
        item = str(row.get("Item", ""))
        qty  = row.get("Qty", None)
        if qty is None:
            vol = row.get("Volume", row.get("VolumeRaw", None))
            qty = extract_numeric_qty(vol) if vol is not None else None
        pdf_items.append({"item": item, "norm": normalize_text(item), "qty": qty})

    # Load workbook (formatting preserved)
    wb = load_workbook(EXCEL_IN)
    if EXCEL_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet '{EXCEL_SHEET}' not found. Sheets: {wb.sheetnames}")
    ws = wb[EXCEL_SHEET]

    # Find headers on HEADER_ROW_NUM
    hdr_row = HEADER_ROW_NUM
    headers = {}
    for col_idx, cell in enumerate(ws[hdr_row], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = col_idx

    if ITEM_HEADER not in headers:
        raise KeyError(f"Header '{ITEM_HEADER}' not found in row {hdr_row}. Found: {list(headers.keys())}")
    if QTY_HEADER not in headers:
        raise KeyError(f"Header '{QTY_HEADER}' not found in row {hdr_row}. Found: {list(headers.keys())}")

    item_col_idx = headers[ITEM_HEADER]
    qty_col_idx  = headers[QTY_HEADER]

    # Build match candidates from sheet rows
    first_data_row = hdr_row + 1
    excel_rows = []
    for r in range(first_data_row, ws.max_row + 1):
        val = ws.cell(row=r, column=item_col_idx).value
        if val is None or str(val).strip() == "":
            continue
        excel_rows.append({"row": r, "text": str(val), "norm": normalize_text(str(val))})

    excel_norm_list = [rec["norm"] for rec in excel_rows]

    # Match each PDF item to best Excel row and write Qty ONLY
    report = []
    written = 0
    for rec in pdf_items:
        if not rec["norm"]:
            report.append({"PDF_Item": rec["item"], "Action": "skipped-empty", "Score": None, "Excel_Item": None})
            continue

        results = process.extract(rec["norm"], excel_norm_list, scorer=fuzz.token_set_ratio, limit=3)
        if not results:
            report.append({"PDF_Item": rec["item"], "Action": "no-candidates", "Score": None, "Excel_Item": None})
            continue

        best_choice, best_score, best_idx = results[0]
        second_score = results[1][1] if len(results) > 1 else None
        delta = (best_score - second_score) if second_score is not None else 999

        match_rec = excel_rows[best_idx]
        excel_text = match_rec["text"]

        if best_score >= MIN_SCORE and delta >= MIN_DELTA and rec["qty"] is not None:
            ws.cell(row=match_rec["row"], column=qty_col_idx).value = rec["qty"]
            action = "write-qty"
            written += 1
        else:
            action = "ambiguous" if second_score is not None else "low-score"

        report.append({
            "PDF_Item": rec["item"],
            "Excel_Item": excel_text,
            "Score": best_score,
            "Second": second_score,
            "Delta": delta if second_score is not None else None,
            "QtyWritten": rec["qty"] if action == "write-qty" else None,
            "Action": action
        })

    # Save BACK to the same file (in place)
    wb.save(EXCEL_IN)

    # Optional: quick console summary
    print(f"✅ Updated in place: {EXCEL_IN}  |  Qty cells written: {written}")
    print(pd.Series([r["Action"] for r in report]).value_counts().to_string())

# ---- call this after you build parsed_df from your PDF ----
# write_qty_in_place(parsed_df)
