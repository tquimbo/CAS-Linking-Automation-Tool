# import re
# from pathlib import Path
# import pandas as pd
# from rapidfuzz import process, fuzz
# from openpyxl import load_workbook

# # ---- Your workbook/config ----
# EXCEL_PATH   = "test_budget.xlsx"
# SHEET_NAME   = "Sheet1"
# ITEM_HEADER  = "Billing Description"  # column header to match against
# COL_RATE_C   = 3   # column C
# COL_QTY_D    = 4   # column D
# REPORT_CSV   = "match_report_from_fast_extract.csv"

# # ---- Import YOUR extractor exactly (no changes) ----
# from fast_extract import parse_via_text, parse_via_ocr, PDF  # uses your constants/funcs

# # same normalizer as earlier (for matching text only; does NOT touch numbers)
# def _norm(s: str) -> str:
#     if not isinstance(s, str):
#         return ""
#     s = s.lower().replace("&", " and ")
#     s = re.sub(r"\[[^\]]*\]", " ", s)
#     s = re.sub(r"[^\w\s]", " ", s)
#     s = re.sub(r"\s+", " ", s).strip()
#     return s

# def _print_like_fast(df: pd.DataFrame, title: str | None = None):
#     if title:
#         print("\n" + title)
#     print(f"{'Item':60} {'Qty':>14} {'Type':>16} {'Rate':>14} {'Total (PDF)':>15} {'Diff':>12}")
#     print("-" * 135)
#     for _, r in df.iterrows():
#         item = (str(r['Item'])[:57] + '...') if len(str(r['Item'])) > 60 else str(r['Item'])
#         qty  = 0.0 if pd.isna(r['Quantity']) else r['Quantity']
#         typ  = r.get('Type') or ''
#         rate = 0.0 if pd.isna(r['Rate']) else r['Rate']
#         tot  = 0.0 if pd.isna(r['Total_PDF']) else r['Total_PDF']
#         diff = 0.0 if pd.isna(r['Diff']) else r['Diff']
#         print(f"{item:60} {qty:14,.4f} {typ:>16} {rate:14,.4f} {tot:15,.2f} {diff:12,.2f}")

# def _get_parsed_df():
#     # follow your fast path + OCR fallback logic exactly
#     df = parse_via_text(PDF)
#     if len(df) < 3:
#         print("Falling back to OCR (this can take a bit)...")
#         df = parse_via_ocr(PDF)
#     return df

# def main():
#     # 1) Parse using YOUR extractor
#     df = _get_parsed_df()
#     if df is None or df.empty:
#         print("No rows parsed with rate & total.")
#         return

#     # 2) Print exactly like your script
#     _print_like_fast(df, title="Parsed from PDF:")

#     # 3) Open workbook (preserve formatting)
#     wb = load_workbook(EXCEL_PATH)
#     if SHEET_NAME not in wb.sheetnames:
#         raise KeyError(f"Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
#     ws = wb[SHEET_NAME]

#     # 4) Locate 'Billing Description' header within top rows and gather rows
#     item_col = None
#     header_row = None
#     for r in range(1, min(ws.max_row, 10) + 1):
#         row_map = {str(c.value).strip(): idx for idx, c in enumerate(ws[r], start=1) if c.value is not None}
#         if ITEM_HEADER in row_map:
#             item_col = row_map[ITEM_HEADER]
#             header_row = r
#             break
#     if item_col is None:
#         raise KeyError(f"Header '{ITEM_HEADER}' not found in first 10 rows.")

#     first_data_row = header_row + 1
#     excel_rows = []
#     for r in range(first_data_row, ws.max_row + 1):
#         desc = ws.cell(row=r, column=item_col).value
#         if desc is None or str(desc).strip() == "":
#             continue
#         cur_rate = ws.cell(row=r, column=COL_RATE_C).value
#         cur_qty  = ws.cell(row=r, column=COL_QTY_D).value
#         excel_rows.append({
#             "row": r,
#             "text": str(desc).strip(),
#             "norm": _norm(str(desc).strip()),
#             "cur_rate": cur_rate,
#             "cur_qty": cur_qty
#         })

#     if not excel_rows:
#         raise RuntimeError("No description rows found under the header to match against.")

#     choices = [e["norm"] for e in excel_rows]

#     # 5) Force-match every parsed row to the closest Excel row; write only if empty
#     report = []
#     wrote_any = 0
#     not_populated_rows = []  # collect parsed rows where we didn't write C or D

#     for idx_df, rec in df.iterrows():
#         src_item = str(rec.get("Item", "")).strip()
#         if not src_item:
#             report.append({"PDF_Item": src_item, "Action": "skipped-empty-item"})
#             not_populated_rows.append(idx_df)
#             continue

#         best, score, idx_choice = process.extractOne(_norm(src_item), choices, scorer=fuzz.token_set_ratio)
#         match = excel_rows[idx_choice]
#         row_idx = match["row"]

#         rate_val = rec.get("Rate", None)
#         qty_val  = rec.get("Quantity", None)  # NOTE: your df uses 'Quantity'

#         wrote_rate = False
#         wrote_qty  = False

#         # only fill empty cells; never overwrite existing values
#         if match["cur_rate"] in (None, "") and pd.notna(rate_val):
#             ws.cell(row=row_idx, column=COL_RATE_C).value = rate_val
#             match["cur_rate"] = rate_val
#             wrote_rate = True

#         if match["cur_qty"] in (None, "") and pd.notna(qty_val):
#             ws.cell(row=row_idx, column=COL_QTY_D).value = qty_val
#             match["cur_qty"] = qty_val
#             wrote_qty = True

#         if wrote_rate or wrote_qty:
#             wrote_any += 1
#             action = "write(C and/or D)"
#         else:
#             action = "skipped-no-empty-targets-or-NaN"
#             not_populated_rows.append(idx_df)

#         report.append({
#             "PDF_Item": src_item,
#             "Excel_Item": match["text"],
#             "Excel_Row": row_idx,
#             "Match_Score": score,
#             "Wrote_Rate(C)": wrote_rate,
#             "Wrote_Qty(D)": wrote_qty,
#             "Existing_Rate(C)_after": match["cur_rate"],
#             "Existing_Qty(D)_after": match["cur_qty"],
#             "Action": action
#         })

#     # 6) Save workbook + audit
#     wb.save(EXCEL_PATH)
#     pd.DataFrame(report).to_csv(REPORT_CSV, index=False)

#     print(f"\n✅ Updated in place: {EXCEL_PATH}")
#     print(f"📝 Match report: {REPORT_CSV}")
#     print(f"Rows where C and/or D was newly filled: {wrote_any} / {len(report)}")

#     # 7) Print items we did NOT populate (same format as initial print)
#     if not_populated_rows:
#         df_unwritten = df.loc[not_populated_rows].copy()
#         _print_like_fast(df_unwritten, title="Items NOT populated into Excel (C/D unchanged):")
#     else:
#         print("\nAll parsed items resulted in a write to column C and/or D.")

# if __name__ == "__main__":
#     main()

import re
from pathlib import Path
import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl import load_workbook

# ---- Your workbook/config ----
EXCEL_PATH   = "test_budget.xlsx"
SHEET_NAME   = "Sheet1"
ITEM_HEADER  = "Billing Description"  # column header to match against
COL_RATE_C   = 3   # column C
COL_QTY_D    = 4   # column D
REPORT_CSV   = "match_report_from_fast_extract.csv"

# ---- Import YOUR extractor exactly (no changes) ----
from fast_extract import parse_via_text, parse_via_ocr, PDF  # uses your constants/funcs

# same normalizer as earlier (for matching text only; does NOT touch numbers)
def _norm(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = s.lower().replace("&", " and ")
    s = re.sub(r"\[[^\]]*\]", " ", s)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _print_like_fast(df: pd.DataFrame, title: str | None = None):
    if title:
        print("\n" + title)
    print(f"{'Item':60} {'Qty':>14} {'Type':>16} {'Rate':>14} {'Total (PDF)':>15} {'Diff':>12}")
    print("-" * 135)
    for _, r in df.iterrows():
        item = (str(r['Item'])[:57] + '...') if len(str(r['Item'])) > 60 else str(r['Item'])
        qty  = 0.0 if pd.isna(r['Quantity']) else r['Quantity']
        typ  = r.get('Type') or ''
        rate = 0.0 if pd.isna(r['Rate']) else r['Rate']
        tot  = 0.0 if pd.isna(r['Total_PDF']) else r['Total_PDF']
        diff = 0.0 if pd.isna(r['Diff']) else r['Diff']
        print(f"{item:60} {qty:14,.4f} {typ:>16} {rate:14,.4f} {tot:15,.2f} {diff:12,.2f}")
    # print totals summary
    print("-" * 135)
    print(f"{'SUM TOTALS':60} {'':14} {'':16} {'':14} "
          f"{df['Total_PDF'].sum(skipna=True):15,.2f} {df['Diff'].sum(skipna=True):12,.2f}")

def _get_parsed_df():
    # follow your fast path + OCR fallback logic exactly
    df = parse_via_text(PDF)
    if len(df) < 3:
        print("Falling back to OCR (this can take a bit)...")
        df = parse_via_ocr(PDF)
    return df

def main():
    # 1) Parse using YOUR extractor
    df = _get_parsed_df()
    if df is None or df.empty:
        print("No rows parsed with rate & total.")
        return

    # 2) Print exactly like your script
    _print_like_fast(df, title="Parsed from PDF:")

    # 3) Open workbook (preserve formatting)
    wb = load_workbook(EXCEL_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    # 4) Locate 'Billing Description' header within top rows and gather rows
    item_col = None
    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        row_map = {str(c.value).strip(): idx for idx, c in enumerate(ws[r], start=1) if c.value is not None}
        if ITEM_HEADER in row_map:
            item_col = row_map[ITEM_HEADER]
            header_row = r
            break
    if item_col is None:
        raise KeyError(f"Header '{ITEM_HEADER}' not found in first 10 rows.")

    first_data_row = header_row + 1
    excel_rows = []
    for r in range(first_data_row, ws.max_row + 1):
        desc = ws.cell(row=r, column=item_col).value
        if desc is None or str(desc).strip() == "":
            continue
        cur_rate = ws.cell(row=r, column=COL_RATE_C).value
        cur_qty  = ws.cell(row=r, column=COL_QTY_D).value
        excel_rows.append({
            "row": r,
            "text": str(desc).strip(),
            "norm": _norm(str(desc).strip()),
            "cur_rate": cur_rate,
            "cur_qty": cur_qty
        })

    if not excel_rows:
        raise RuntimeError("No description rows found under the header to match against.")

    choices = [e["norm"] for e in excel_rows]

    # 5) Force-match every parsed row to the closest Excel row; write only if empty
    report = []
    wrote_any = 0
    not_populated_rows = []  # collect parsed rows where we didn't write C or D

    for idx_df, rec in df.iterrows():
        src_item = str(rec.get("Item", "")).strip()
        if not src_item:
            report.append({"PDF_Item": src_item, "Action": "skipped-empty-item"})
            not_populated_rows.append(idx_df)
            continue

        best, score, idx_choice = process.extractOne(_norm(src_item), choices, scorer=fuzz.token_set_ratio)
        match = excel_rows[idx_choice]
        row_idx = match["row"]

        rate_val = rec.get("Rate", None)
        qty_val  = rec.get("Quantity", None)  # NOTE: your df uses 'Quantity'

        wrote_rate = False
        wrote_qty  = False

        # only fill empty cells; never overwrite existing values
        if match["cur_rate"] in (None, "") and pd.notna(rate_val):
            ws.cell(row=row_idx, column=COL_RATE_C).value = rate_val
            match["cur_rate"] = rate_val
            wrote_rate = True

        if match["cur_qty"] in (None, "") and pd.notna(qty_val):
            ws.cell(row=row_idx, column=COL_QTY_D).value = qty_val
            match["cur_qty"] = qty_val
            wrote_qty = True

        if wrote_rate or wrote_qty:
            wrote_any += 1
            action = "write(C and/or D)"
        else:
            action = "skipped-no-empty-targets-or-NaN"
            not_populated_rows.append(idx_df)

        report.append({
            "PDF_Item": src_item,
            "Excel_Item": match["text"],
            "Excel_Row": row_idx,
            "Match_Score": score,
            "Wrote_Rate(C)": wrote_rate,
            "Wrote_Qty(D)": wrote_qty,
            "Existing_Rate(C)_after": match["cur_rate"],
            "Existing_Qty(D)_after": match["cur_qty"],
            "Action": action
        })

    # 6) Save workbook + audit
    wb.save(EXCEL_PATH)
    pd.DataFrame(report).to_csv(REPORT_CSV, index=False)

    print(f"\n✅ Updated in place: {EXCEL_PATH}")
    print(f"📝 Match report: {REPORT_CSV}")
    print(f"Rows where C and/or D was newly filled: {wrote_any} / {len(report)}")

    # 7) Print items we did NOT populate (same format as initial print)
    if not_populated_rows:
        df_unwritten = df.loc[not_populated_rows].copy()
        _print_like_fast(df_unwritten, title="Items NOT populated into Excel (C/D unchanged):")
        # also print sum of their totals
        print(f"\n⚠️  Sum of Total_PDF for NOT populated items: {df_unwritten['Total_PDF'].sum(skipna=True):,.2f}")
    else:
        print("\nAll parsed items resulted in a write to column C and/or D.")

    # 8) Print final grand totals for full dataset
    print("\n📊 GRAND TOTALS:")
    print(f"   Parsed Total_PDF Sum: {df['Total_PDF'].sum(skipna=True):,.2f}")
    print(f"   Parsed Diff Sum:       {df['Diff'].sum(skipna=True):,.2f}")

if __name__ == "__main__":
    main()
